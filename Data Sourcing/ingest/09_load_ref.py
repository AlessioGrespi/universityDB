#!/usr/bin/env python3
"""Load REF 2021 results from XLSX."""

import sys
from openpyxl import load_workbook
from config import RAW_DIR
from db import get_conn, fetch_all, bulk_insert, execute
from utils import UniversityMatcher


def main():
    src = RAW_DIR / "ref2021_results.xlsx"
    if not src.exists():
        print(f"Run fetch_all_datasets.py first — {src} not found")
        sys.exit(1)

    print("Loading REF 2021 results...")
    wb = load_workbook(src, data_only=True)
    ws = wb.active

    # Header is on row 7 (0-indexed row 6), data starts row 8
    rows_iter = ws.iter_rows(values_only=True)
    # Skip to header row
    for i in range(7):
        row = next(rows_iter)
        if i == 6:
            headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
    print(f"  Columns: {headers}")

    data = []
    for row in rows_iter:
        d = dict(zip(headers, row))
        if d.get("Institution name"):
            data.append(d)

    print(f"  {len(data)} data rows")

    with get_conn() as conn:
        unis = fetch_all(conn, "SELECT id, name, ukprn, hesa_id FROM universities")
        matcher = UniversityMatcher(unis)

        execute(conn, "DELETE FROM ref_results")

        batch = []
        skipped = 0
        for d in data:
            inst_name = d.get("Institution name", "")
            inst_ukprn = str(d.get("Institution code (UKPRN)", "")).strip()
            uni_id = matcher.match(name=inst_name, ukprn=inst_ukprn)
            if not uni_id:
                skipped += 1
                continue

            def safe_float(v):
                try:
                    return float(v) if v is not None else None
                except (ValueError, TypeError):
                    return None

            def safe_int(v):
                try:
                    return int(v) if v is not None else None
                except (ValueError, TypeError):
                    return None

            uoa_name = d.get("Unit of assessment name", "")
            uoa_num = safe_int(d.get("Unit of assessment number"))
            profile = d.get("Profile", "")
            fte = safe_float(d.get("FTE of submitted staff"))

            batch.append((
                uni_id,
                uoa_name,
                uoa_num,
                profile,
                fte,
                safe_float(d.get("4*")),
                safe_float(d.get("3*")),
                safe_float(d.get("2*")),
                safe_float(d.get("1*")),
                safe_float(d.get("Unclassified")),
            ))

        if batch:
            bulk_insert(
                conn, "ref_results",
                ["university_id", "unit_of_assessment", "uoa_number", "profile_type",
                 "staff_fte", "pct_4star", "pct_3star", "pct_2star", "pct_1star", "pct_unclassified"],
                batch,
            )

    print(f"  Done — {len(batch)} rows loaded ({skipped} skipped — institution not matched)")


if __name__ == "__main__":
    main()
